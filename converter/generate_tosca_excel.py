"""
Tosca Excel Generator - Most Compatible Import Method
Generates Excel files compatible with Tosca's built-in Excel import
Works with ALL Tosca versions (9.0+)
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Any

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not installed. Install with: uv pip install openpyxl")
    EXCEL_AVAILABLE = False


class ToscaExcelGenerator:
    """
    Generates Tosca-compatible Excel files from migration.json
    Compatible with Tosca's built-in Excel import feature
    """
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: uv pip install openpyxl")
        
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
    
    def generate_from_migration(self, migration_json_path: str, output_folder: str) -> Dict[str, Any]:
        """Generate Tosca Excel files from migration.json"""
        logger.info(f"Loading migration file: {migration_json_path}")
        
        with open(migration_json_path, 'r', encoding='utf-8') as f:
            migration_data = json.load(f)
        
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)
        
        stats = {
            "modules_created": 0,
            "testcases_created": 0,
            "files_generated": []
        }
        
        # Generate Modules Excel
        if "objects" in migration_data:
            module_file = output_path / "Modules.xlsx"
            self._generate_modules_excel(migration_data, module_file)
            stats["modules_created"] = 1
            stats["files_generated"].append(str(module_file))
            logger.info(f"✓ Generated: {module_file}")
        
        # Generate TestCases Excel  
        if "testCases" in migration_data:
            testcase_file = output_path / "TestCases.xlsx"
            self._generate_testcases_excel(migration_data, testcase_file)
            stats["testcases_created"] = len(migration_data["testCases"])
            stats["files_generated"].append(str(testcase_file))
            logger.info(f"✓ Generated: {testcase_file}")
        
        return stats
    
    def _generate_modules_excel(self, migration_data: Dict, output_file: Path):
        """Generate Modules Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modules"
        
        # Header row
        headers = ["Module Name", "Control Name", "Technique", "Selector", "Control Type", "Engine"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        module_name = migration_data.get("metadata", {}).get("projectName", "CenteneHomePage")
        objects = migration_data.get("objects", [])
        
        for obj in objects:
            selector = obj.get("selector", "")
            
            # Convert Cypress selectors to XPath if needed
            if ":contains(" in selector:
                selector = self._convert_selector_to_xpath(selector)
            
            ws.append([
                module_name,
                obj.get("name", "Unknown"),
                obj.get("technique", "XPath"),
                selector,
                self._map_control_type(obj.get("name", "")),
                "TBox Web"
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        logger.info(f"Created module Excel with {len(objects)} controls")
    
    def _generate_testcases_excel(self, migration_data: Dict, output_file: Path):
        """Generate TestCases Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestCases"
        
        # Header row
        headers = ["TestCase Name", "Step #", "Action", "Module", "Control", "Value", "Description"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        testcases = migration_data.get("testCases", [])
        module_name = migration_data.get("metadata", {}).get("projectName", "CenteneHomePage")
        
        for tc_data in testcases:
            tc_name = tc_data.get("name", "Untitled")
            steps = tc_data.get("steps", [])
            
            for idx, step_data in enumerate(steps, 1):
                action = self._map_action(step_data.get("action", ""))
                control = step_data.get("target", "")
                value = step_data.get("value", step_data.get("url", ""))
                
                ws.append([
                    tc_name,
                    idx,
                    action,
                    module_name if control else "",
                    control,
                    value,
                    step_data.get("notes", "")
                ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        logger.info(f"Created {len(testcases)} test cases in Excel")
    
    def _convert_selector_to_xpath(self, selector: str) -> str:
        """Convert Cypress selectors to XPath"""
        import re
        
        # button:contains('text') → //button[contains(text(),'text')]
        pattern = r"([a-zA-Z]+):contains\(['\"]([^'\"]+)['\"]\)"
        match = re.search(pattern, selector)
        if match:
            return f"//{match.group(1)}[contains(text(),'{match.group(2)}')]"
        
        # :contains('text') → //*[contains(text(),'text')]
        pattern2 = r":contains\(['\"]([^'\"]+)['\"]\)"
        match2 = re.search(pattern2, selector)
        if match2:
            return f"//*[contains(text(),'{match2.group(1)}')]"
        
        return selector
    
    def _map_action(self, generic_action: str) -> str:
        """Map generic action to Tosca ActionWord"""
        action_map = {
            "Navigate": "TBox Navigate",
            "Click": "TBox Click",
            "Input": "TBox Enter Text",
            "Verify": "TBox Verify Attribute",
            "VerifyText": "TBox Verify Text",
            "Wait": "TBox Wait"
        }
        return action_map.get(generic_action, generic_action)
    
    def _map_control_type(self, name: str) -> str:
        """Infer control type from name"""
        name_lower = name.lower()
        if "btn" in name_lower or "button" in name_lower:
            return "Button"
        elif "input" in name_lower:
            return "TextBox"
        elif "link" in name_lower:
            return "Link"
        elif "dropdown" in name_lower:
            return "ComboBox"
        else:
            return "Control"


def main():
    """Command-line interface"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Tosca Excel files from migration.json")
    parser.add_argument("migration_json", help="Path to migration.json file")
    parser.add_argument("--output", help="Output folder for Excel files", default="tosca-output/excel")
    
    args = parser.parse_args()
    
    generator = ToscaExcelGenerator()
    stats = generator.generate_from_migration(args.migration_json, args.output)
    
    print("\n" + "="*80)
    print("✅ EXCEL GENERATION COMPLETE")
    print("="*80)
    print(f"\nModules Created: {stats['modules_created']}")
    print(f"TestCases Created: {stats['testcases_created']}")
    print(f"\nFiles Generated:")
    for file in stats['files_generated']:
        print(f"  • {file}")
    print("\n" + "="*80)
    print("HOW TO IMPORT:")
    print("1. Open Tosca Commander")
    print("2. Go to: TOOLS → Import → Import from Excel")
    print("3. Select: Modules.xlsx")
    print("4. Follow the import wizard")
    print("5. Repeat for: TestCases.xlsx")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
